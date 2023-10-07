from openpyxl import *
from tkinter import *
# tạo biến đường dẫn lưu file excel hiện có
wb = load_workbook('E:\\PY\\Lab04\\cau05.xlsx')
# tạo đối tượng sheet
sheet = wb.active
def excel():
    # đặt lại kích thương cột trong sheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50
    # điền dữ liệu vào vị trí cụ thể
    sheet.cell(row=1,column=1 ).value = "Name"
    sheet.cell(row=1,column=2 ).value = "Course"
    sheet.cell(row=1,column=3 ).value = "Semester"
    sheet.cell(row=1,column=4 ).value = "Form Number"
    sheet.cell(row=1,column=5 ).value = "Contact Number"
    sheet.cell(row=1,column=6 ).value = "Email id"
    sheet.cell(row=1,column=7 ).value = "Address"
# hàm khi thực hiện sự kiện chọn (con trỏ)
def focus1(event):
    # thiết lập chọn hộp course_field 
    course_field.focus_set()

# hàm khi thực hiện sự kiện chọn
def focus2(event):
    # thiết lập chọn hộp sem_field
    sem_field.focus_set()
# hàm khi thực hiện sự kiện chọn
def focus3(event):
    # thiết lập chọn hộp form_no_field
    form_no_field.focus_set()
# hàm khi thực hiện sự kiện chọn
def focus4(event):
    # thiết lập chọn hộp contact_no_field
    contact_no_field.focus_set()
# hàm khi thực hiện sự kiện chọn
def focus5(event):
    #Thiết lập chọn hộp email_id_field
    email_id_field.focus_set()
# hàm khi thực hiện sự kiện chọn
def focus6(event):
    #thiết lập chọn hộp address_field
    address_field.focus_set()

# hàm xóa nội dung trong hộp 
def clear():
    # xóa nội dung trong hộp
    name_field.delete(0,END)
    course_field.delete(0,END)
    sem_field.delete(0,END)
    form_no_field.delete(0,END)
    contact_no_field.delete(0,END)
    email_id_field.delete(0,END)
    address_field.delete(0,END)

#Hàm thêm nội dung từ giao diện
def insert():
    # nếu nội dung trống thông báo người dùng hãy nhập dữ liệu
    if (name_field.get() =="" and
        course_field.get()==""and
        sem_field.get()==""and
        form_no_field.get()==""and
        contact_no_field.get()==""and
        email_id_field.get()==""and
        address_field.get()==""):

        print("hãy nhập dữ liệu")
    else:
        # gán hàng tối đa và cột tối đa, gán giá trị tối đa cho dữ liệu được ghi tới biến
        current_row = sheet.max_row
        current_column = sheet.max_column

        # phương thức get trả về văn bản hiện tại dưới dạng chuỗi 
        # được viết vào vị trí cụ thể trên sheet
        sheet.cell(row=current_row+1,column=1).value =name_field.get()
        sheet.cell(row=current_row+1,column=2).value =course_field.get()
        sheet.cell(row=current_row+1,column=3).value =sem_field.get()
        sheet.cell(row=current_row+1,column=4).value =form_no_field.get()
        sheet.cell(row=current_row+1,column=5).value =contact_no_field.get()
        sheet.cell(row=current_row+1,column=6).value =email_id_field.get()
        sheet.cell(row=current_row+1,column=7).value =address_field.get()

        # lưu vào địa chỉ file
        wb.save('E:\\PY\\Lab04\\cau05.xlsx')  

        # thiết lập chọn hộp name_field
        name_field.focus_set()
        # gọi hàm clear()     
        clear()
if __name__=="__main__":
    # tạo giao cửa sổ giao diện người dùng
    root=Tk()

    # thiết lập nền cho cửa sổ giao diện người dùng
    root.configure(background='light green')

     # thiết lập tiêu đề cho cửa sổ giao diện người dùng
    root.title("registration form")

    # thiết lập cấu hình cho cửa sổ giao diện người dùng
    root.geometry("500x300")

    excel()

    # tạo các trường liên quan
    heading = Label(root,text="Form",bg="light green")
    name =Label(root,text="Name",bg="light green")
    course = Label(root,text="Course",bg="light green")
    sem = Label(root,text="Semester",bg="light green")
    form_no =Label(root,text="Form No.",bg="light green")
    contact_no =Label(root,text="Contact No.",bg="light green")
    email_id =Label(root,text="Email id",bg="light green")
    address = Label(root,text="Address",bg="light green")

    # sử dụng grid để đặt các vị trí control
    heading.grid(row=0,column=1)
    name.grid(row=1,column=0)
    course.grid(row=2,column=0)
    sem.grid(row=3,column=0)
    form_no.grid(row=4,column=0)
    contact_no.grid(row=5,column=0)
    email_id.grid(row=6,column=0)
    address.grid(row=7,column=0)

    # tạo các Entry textbox
    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)

    # ràng buộc các widget (control) với sự kiện
    # khi nhấn nút enter gọi sự kiện
    name_field.bind("<Return>",focus1)
    course_field.bind("<Return>",focus2)
    sem_field.bind("<Return>",focus3)
    form_no_field.bind("<Return>",focus4)
    contact_no_field.bind("<Return>",focus5)
    email_id_field.bind("<Return>",focus6)

    # sử dụng grid để đưa các Entry textbox lên giống dạng bảng
    name_field.grid(row=1,column=1,ipadx="100")
    course_field.grid(row=2,column=1,ipadx="100")
    sem_field.grid(row=3,column=1,ipadx="100")
    form_no_field.grid(row=4,column=1,ipadx="100")
    contact_no_field.grid(row=5,column=1,ipadx="100")
    email_id_field.grid(row=6,column=1,ipadx="100")
    address_field.grid(row=7,column=1,ipadx="100")

    # gọi hàm excel
    excel()

    # tạo nút vào đặt vào trong cửa sổ root
    submit = Button(root,text="Submit",fg="Black",bg="Red",command=insert)
    # đưa nút Submit lên cửa sổ giao diện người dùng
    submit.grid(row=8,column=1)

    # chạy giao diện người dùng
    root.mainloop()