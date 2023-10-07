from tkinter import messagebox
from openpyxl import *
from tkinter import * 
import tkinter as tk
from tkinter.ttk import Combobox 
from datetime import datetime
import re

# Khởi tạo đường dẫn lưu file
wb = load_workbook('E:\\PY\\Lab04\\cau06.xlsx')
sheet = wb.active

def excel():
    sheet.column_dimensions['A'].width=10
    sheet.column_dimensions['B'].width=30
    sheet.column_dimensions['C'].width=10
    sheet.column_dimensions['D'].width=20
    sheet.column_dimensions['E'].width=10
    sheet.column_dimensions['F'].width=10
    sheet.column_dimensions['G'].width=10
    sheet.column_dimensions['H'].width=50

    sheet.cell(row=1,column=1).value="Mã số sinh viên"
    sheet.cell(row=1,column=2).value="Họ tên"
    sheet.cell(row=1,column=3).value="Ngày sinh"
    sheet.cell(row=1,column=4).value="Email"
    sheet.cell(row=1,column=5).value="Số điện thoại"
    sheet.cell(row=1,column=6).value="Học kỳ"
    sheet.cell(row=1,column=7).value="Năm học"
    sheet.cell(row=1,column=8).value="Môn học"
def focus1(envet):
    txtMs.focus_set()
def focus2(event):
    txtHoTen.focus_set()
def focus3(event):
    txtNgaySinh.focus_set()
def focus4(event):
    txtEmail.focus_set()
def focus5(event):
    txtSDT.focus_set()
def focus6(event):
    txtHocKy.focus_set()
def focus7(event):
    cboNamHoc.focus_set()

def clear():
    txtMs.delete(0,END)
    txtHoTen.delete(0,END)
    txtNgaySinh.delete(0,END)
    txtSDT.delete(0,END)
    txtHocKy.delete(0,END)
    cboNamHoc.delete(0,END)
    txtEmail.delete(0,END)
    checkbox1.option_clear()
    checkbox2.option_clear()
    checkbox3.option_clear()
    checkbox4.option_clear()
def insert():
    if (txtMs.get()==""and
        txtHoTen.get()==""and
        txtNgaySinh.get()==""and
        txtEmail.get()==""and
        txtSDT.get()==""and
        txtHocKy.get()==""and
        cboNamHoc.get()==""and
        ckb1.get()==0 and
        ckb2.get()==0 and
        ckb3.get()==0 and
        ckb4.get()==0):

        print("Hay nhập dữ liệu")
    else:
        
        _row = sheet.max_row
        _column=sheet.max_column
        check=0
        kq = ""
        if ckb1.get()==1:
            kq+= "Lập trình Python  "
            check+=1
        if ckb2.get()==1:
            kq+= "Lập trình Java  "
            check+=1
        if ckb3.get()==1:
            kq+= "Công nghệ phần mềm  "
            check+=1
        if ckb4.get()==1:
            kq+= "Phát triển ứng dụng web  "
            check+=1
       
        sheet.cell(row=_row+1,column=1).value=txtMs.get()
        sheet.cell(row=_row+1,column=2).value=txtHoTen.get()
        sheet.cell(row=_row+1,column=3).value=txtNgaySinh.get()
        sheet.cell(row=_row+1,column=4).value=txtEmail.get()
        sheet.cell(row=_row+1,column=5).value=txtSDT.get()
        sheet.cell(row=_row+1,column=6).value=txtHocKy.get()
        sheet.cell(row=_row+1,column=7).value=cboNamHoc.get()
        sheet.cell(row=_row+1,column=8).value=kq

        kiemTra(check)

def kiemTra(check:int):
    ms = txtMs.get()
    mail=txtEmail.get()
    Sdt=txtSDT.get()
    hk=txtHocKy.get()
    ns=txtNgaySinh.get()
    mau_ngay = "%d/%m/%Y"
    mau_mail = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    

    if len(ms)==7  and len(Sdt)==10:
        if Sdt.isnumeric():
            if hk == '1' or hk == '2' or hk =='3':  
                if re.match(mau_mail,mail):
                    try:
                        datetime.strptime(ns,mau_ngay)
                        if check>0:
                            wb.save('E:\\PY\\Lab04\\cau06.xlsx')
                            txtMs.focus_set()
                            messagebox.showinfo("Thông báo",f"Đã thêm thành công sinh viên có mã {ms}"  )
                            clear()
                        else:
                             messagebox.showerror("Thông báo","Chưa đăng ký môn học nào")
                    except ValueError:
                        messagebox.showerror("Thông báo","vui lòng kiểm tra định dạng ngày tháng năm")
            else:
                messagebox.showerror("Thông báo","vui lòng nhập học kì 1 hoặc 2 hoặc 3")
        else:
            messagebox.showerror("Thông báo","Lỗi tại ô số điện thoại")
    else:
        messagebox.showerror("Thông báo","hãy nhập đúng mã số hoặc số điện thoại")

    


# kiểm tra file đang chạy
if __name__=="__main__":
    # khởi tạo GUI
    root = Tk()
    # thiết lập nền cho GUI
    root.configure(background="light green")
    # thiết lập tiêu để cho ứng dụng
    root.title("Đăng ký học phần")
    # thiết lập cấu hình cửa số
    root.geometry("600x300")

    excel()

    # thiết lập các label liên quan
    TieuDe=Label(root,text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN",bg="light green",fg="Red",font=("",16))
    mssv=Label(root,text="Mã số sinh viên",bg="light green")
    hoTen=Label(root,text="Họ Tên",bg="light green")
    ngaySinh=Label(root,text="Ngày Sinh",bg="light green")
    email=Label(root,text="Email",bg="light green")
    sdt=Label(root,text="Số điện thoại",bg="light green")
    hocKy=Label(root,text="Học Kỳ",bg="light green")
    namHoc=Label(root,text="Năm Học",bg="light green")
    chonMonHoc=Label(root,text="Chọn môn học",bg="light green")
    # thiết lập các ô nhập liệu liên quan
    txtMs=Entry(root)
    txtHoTen=Entry(root)
    txtNgaySinh=Entry(root)
    txtEmail=Entry(root)
    txtSDT=Entry(root)
    txtHocKy=Entry(root)
    cboNamHoc= Combobox(root,values=["2021-2022","2022-2023","2023-2024","2024-2025"])
    ## tạo checkbox gán giá trị
    ckb1=IntVar()
    ckb2=IntVar()
    ckb3=IntVar()
    ckb4=IntVar()
    checkbox1 = tk.Checkbutton(root,text="Lập trình Python",variable=ckb1,bg="light green")
    checkbox2 = tk.Checkbutton(root,text="Lập trình Java",variable=ckb2,bg="light green")
    checkbox3 = tk.Checkbutton(root,text="Công nghệ phần mềm",variable=ckb3,bg="light green")
    checkbox4 = tk.Checkbutton(root,text="Phát triển ứng dụng web",variable=ckb4,bg="light green")

    txtMs.bind("<Return>",focus1)
    txtHoTen.bind("<Return>",focus2)
    txtNgaySinh.bind("<Return>",focus3)
    txtEmail.bind("<Return>",focus4)
    txtSDT.bind("<Return>",focus5)
    txtHocKy.bind("<Return>",focus6)
    cboNamHoc.bind("<Return>",focus7)

    # Thêm các widget lên GUI
    TieuDe.grid(row=0,column=1)

    mssv.grid(row=1,column=0,sticky=W)
    txtMs.grid(row=1,column=1,ipadx="100")

    hoTen.grid(row=2,column=0,sticky=W)
    txtHoTen.grid(row=2,column=1,ipadx="100")

    ngaySinh.grid(row=3,column=0,sticky=W)
    txtNgaySinh.grid(row=3,column=1,ipadx="100")

    email.grid(row=4,column=0,sticky=W)
    txtEmail.grid(row=4,column=1,ipadx="100")

    sdt.grid(row=5,column=0,sticky=W)
    txtSDT.grid(row=5,column=1,ipadx="100")

    hocKy.grid(row=6,column=0,sticky=W)
    txtHocKy.grid(row=6,column=1,ipadx="100")

    namHoc.grid(row=7,column=0,sticky=W)
    cboNamHoc.grid(row=7,column=1,ipadx="90")

    chonMonHoc.grid(row=8,column=0,sticky=W)
    checkbox1.grid(row=9,column=1,sticky=W)
    checkbox2.grid(row=9,column=2,sticky=W)
    checkbox3.grid(row=10,column=1,sticky=W)
    checkbox4.grid(row=10,column=2,sticky=W)

    excel()

    dangKy = Button(root,text="Đăng Ký",fg="Black",bg="Red",command=insert)
    Thoat =Button(root,text="Thoát",fg="Black",bg="Red",command=exit)
    dangKy.grid(row=11,column=1)
    Thoat.grid(row=11,column=2)


    
    root=mainloop()
#--------------------------------------
